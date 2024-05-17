import sys

module_location = r"C:\FLESTIM\src"
sys.path.append(module_location)

from openpyxl import load_workbook
from player.player import *

class data:
    FIRST_ROW = 8
    FIRST_COLONNE = 1
    LAST_COLONNE = 5

    FIRST_COLONNE_ROUGE = 1
    FIRST_COLONNE_BLEU = 8


    def __init__(self, chemin, nbEquipe, equipeBleu,equipeRouge):
        self.equipeBleu = equipeBleu
        self.equipeRouge = equipeRouge

        self.cheminExcel = chemin+r"\nomenclature.xlsx"
        
        # self.cheminExcel = r"C:\FLESTIM\nomenclature.xlsx"
        self.fichierExcel = load_workbook(filename=self.cheminExcel, read_only=True,data_only=True)
        self.feuille = self.fichierExcel.active

        # Lire cellule -> Nombre de VM pour chaque
        self.nbEquipe = nbEquipe
  
        self.extraireDonnees()
    
    def obtenirLigne (self,index, debutColonne):
        return [self.feuille.cell(row=index+self.FIRST_ROW, column=debutColonne + i).value for i in range(self.LAST_COLONNE)]
    
    def obtenirUserCible (self,index):
        return self.feuille.cell(row=index+self.FIRST_ROW, column=14).value
    
    def extraireDonnees(self): 
        # tableau = self.feuille.iter_rows(min_row=data.FIRST_ROW, min_col=(data.FIRST_COLONNE + debutColonne), max_col=(data.LAST_COLONNE + debutColonne),max_row=(nbVM+data.FIRST_ROW))
        for index in range(self.nbEquipe):
            listeBleu = self.obtenirLigne(index,self.FIRST_COLONNE_BLEU)
            listeRouge = self.obtenirLigne(index,self.FIRST_COLONNE_ROUGE)
            userCible = self.obtenirUserCible(index)
            
            joueurBleu = playerVictime(listeBleu,self.equipeBleu,userCible)
            self.equipeBleu.addPlayer(joueurBleu)

            joueurRouge = playerAttaque(listeRouge,self.equipeRouge,joueurBleu)
            self.equipeRouge.addPlayer(joueurRouge)

            
        self.fichierExcel.close()

