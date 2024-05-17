import os

from openpyxl import load_workbook
from src.player.player import *

class traitementDonnee:
    NOM_EXCEL = r"nomenclature.xlsx"
    FIRST_ROW = 8
    FIRST_COLONNE = 1
    LAST_COLONNE = 5

    FIRST_COLONNE_ROUGE = 1
    FIRST_COLONNE_BLEU = 8


    def __init__(self, chemin_module, nbEquipe, equipeBleu,equipeRouge):
        # repertoire = os.path.dirname(__file__)

        self.equipeBleu = equipeBleu
        self.equipeRouge = equipeRouge

        self.cheminExcel = self.NOM_EXCEL
        
        self.fichierExcel = load_workbook(filename=self.cheminExcel, read_only=True,data_only=True)
        self.feuille = self.fichierExcel.active

        self.nbEquipe = nbEquipe
  
        self.extraireDonnees()
    
    def obtenirLigne (self,index, debutColonne):
        return [self.feuille.cell(row=index+self.FIRST_ROW, column=debutColonne + i).value for i in range(self.LAST_COLONNE)]
    
    def obtenirUserCible (self,index):
        return self.feuille.cell(row=index+self.FIRST_ROW, column=14).value
    
    def extraireDonnees(self): 
        for index in range(self.nbEquipe):
            listeBleu = self.obtenirLigne(index,self.FIRST_COLONNE_BLEU)
            listeRouge = self.obtenirLigne(index,self.FIRST_COLONNE_ROUGE)
            userCible = self.obtenirUserCible(index)
            
            joueurBleu = playerVictime(listeBleu,self.equipeBleu,userCible)
            self.equipeBleu.addPlayer(joueurBleu)

            joueurRouge = playerAttaque(listeRouge,self.equipeRouge,joueurBleu)
            self.equipeRouge.addPlayer(joueurRouge)

            
        self.fichierExcel.close()

